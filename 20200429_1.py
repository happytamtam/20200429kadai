import openpyxl
import pandas as pd

#課題2
df = pd.read_csv('C:/Users/hiro/Desktop/fruit.csv', encoding='cp932')
print(df)

#課題3
exel_file = "C:/Users/hiro/Desktop/sample.xlsx"
#3-1
wb = openpyxl.load_workbook(exel_file)

#3-3シートの取得
ws = wb.worksheets[0]
print(ws)

#3-4シート名変更
ws.title = "SheetOne"
sheetlists = wb.sheetnames
print(sheetlists)

#3-5シート追加
ws4 = wb.create_sheet(title="Sheet4")
sheetlists = wb.sheetnames
print(sheetlists)

#3-6シートコピー
ws2_copy = wb.copy_worksheet(wb["Sheet2"])
sheetlists = wb.sheetnames
print(sheetlists)

#3-7シート削除
#wb.remove(ws2_copy)
sheetlists = wb.sheetnames[4:]
for sheet in sheetlists:
    wb.remove(wb[sheet])

sheetlists = wb.sheetnames
print(sheetlists)

#3-8セルの取得
c1 = ws.cell(1, 1)
print(c1)

#3-9行の取得
row1 = ws[1]
print(row1)

#3-10範囲取得
rng1 = ws["A1:A5"]
print(rng1)

#3-11値読み込み
ws1 = wb["SheetOne"]
val1 = ws1.cell(1,1).value
print(val1)

#3-12シート名全て取得
for sheets in wb:
    print(sheets.title)

#3-13シート内の値を取得
addrs = []
for row in ws.rows:
    for cell in row:
        addrs.append(cell.value)
print(addrs)

#3-14範囲指定して値を取得
addrs2 = []
for row in ws.iter_rows(min_row=2, min_col=2):
	for cell in row:
		addrs2.append(cell.value)
print(addrs2)

#3-2
wb.save(exel_file)



